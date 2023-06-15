import PyPDF2
import re
import pandas as pd
from openpyxl import load_workbook

def extract(pdfs):
    with open(pdfs, "rb") as pdf:
        reader = PyPDF2.PdfReader(pdf, strict=False)
        text = []

        for page in reader.pages:
            content = page.extract_text()
            text.append(content)

        return text
#AC-Catalog-2022-2023-08082022
extractedText = extract('completed colleges/andrew.pdf')

book = load_workbook('data.xlsx')
sheet=book.worksheets[0]



df = pd.DataFrame(columns=['Colleges','Majors','Courses'])
pattern = r"[A-Z]{3} [0-9]{3} - [A-Z].+\("
majors = r"[A-Z]{3} - [A-Za-z].+?\."
majormatches = []
matches = []

for maj in extractedText:
    mat = re.findall(majors, maj)
    if mat:
        majormatches.extend(mat)

for text in extractedText:
    match = re.findall(pattern, text)
    if match:
        matches.extend(match)

count = 0
for mm in majormatches:
    mm = mm.strip(" .")
    if mm == "ACE - Andrew College Experience":
        continue
    majorname = mm[0:3]
    print('---------------------------------')
    mm = mm[6:]
    print(mm)
    for m in matches:
        if count == 0 and m == "CRM 220 - Clinical Practicum (":
            count += 1
            continue
        m = m.strip("(")
        if m[0:3] == majorname:
            print(m)
            df.loc[len(df.index)] = ["Andrew College",majorname, m]
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)