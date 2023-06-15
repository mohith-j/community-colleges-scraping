import PyPDF2
import re
import pandas as pd
from openpyxl import load_workbook

def extract(pdfs):
	with open(pdfs, "rb") as pdf:
		reader = PyPDF2.PdfReader(pdf, strict=False)
		text = []

		for page in reader.pages:
			content = page.extract_text()
			text.append(content)

		return text

extractedText = extract('completed colleges/southernregional.pdf')

book = load_workbook('data.xlsx')
sheet=book.worksheets[0]



df = pd.DataFrame(columns=['Colleges','Majors','Courses'])

pattern = r"[A-Z]{4} \d{4} \- [A-Z].+"
matches = []

for text in extractedText:
	match = re.findall(pattern, text)
	if match:
		matches.extend(match)


li = []
final = []
for m in matches:
	if m[0:9] not in li:
		li.append(m[0:9])
		final.append(m)
final = sorted(final)
currmjr = ""
for i in final:
	if currmjr != i[0:4]:
		print('---------------------------------')
		print(i[0:4])
	currmjr = i[0:4]
	print(i)
	# if "ACCT 2115 - Bookkeeper Certication Review" in i:
	# 	# df.loc[len(df.index)] = ["Southern Regional Technical College",currmjr, "ACCT 2115 - Bookkeeper Certication Review"]
	# 	print("wassup")
	df.loc[len(df.index)] = ["Southern Regional Technical College",currmjr, i]
df = df.applymap(lambda x: x.encode('unicode_escape').
                 decode('utf-8') if isinstance(x, str) else x)
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)