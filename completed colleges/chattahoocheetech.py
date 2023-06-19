from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl import load_workbook

def namechange(name):
	link=name.replace(" - ","-")
	link = link.replace("- ", "-")
	link = link.replace("/", "-")
	link =link.replace(" ", "-")
	link="https://chattahoocheetech.smartcatalogiq.com/en/2022-2023/General-Catalog/Courses/"+link
	if "CMTT" in name:
		link="https://chattahoocheetech.smartcatalogiq.com/2022-2023/General-Catalog/Courses/CMMT"
	if "DMPT" in name:
		link = "https://chattahoocheetech.smartcatalogiq.com/2022-2023/General-Catalog/Courses/DMPT-Design-and-Media-Production"
	if "ELUT" in name:
		link = "https://chattahoocheetech.smartcatalogiq.com/2022-2023/General-Catalog/Courses/ELUT"
	r = requests.get(link)
	if r.status_code != 404:
		return link
	else:
		return None


df = pd.DataFrame(columns=['Colleges','Majors','Courses'])
url = "https://chattahoocheetech.smartcatalogiq.com/en/2022-2023/General-Catalog/Courses"
html = requests.get(url)
soup = BeautifulSoup(html.text, "html.parser")
majors = soup.findAll("a", attrs={"class":None}, href=True)
book = load_workbook('data.xlsx')
sheet=book.worksheets[0]

for major in majors:
	if "General Catalog" in major.text:
		continue
	if "-" in major.text:
		print('---------------------------------')
		print(major.text)
		newurl = namechange(major.text)
		newhtml = requests.get(newurl)
		newsoup = BeautifulSoup(newhtml.text, "html.parser")
		classes = newsoup.findAll("a", attrs={"class":None}, href=True)
		for classy in classes:
			if "General Catalog" in classy.text:
				continue
			if ("0" in classy.text or "1" in classy.text or "2" in classy.text) and len(classy.text) > 5:
				print(classy.text)
				df.loc[len(df.index)] = ["Chattahoochee Technical College",major.text, classy.text]
		if "Welding" in major.text:
			break
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)
 