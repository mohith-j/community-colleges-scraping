import PyPDF2
import re
import pandas as pd
from openpyxl import load_workbook

def extract(pdfs):
    with open(pdfs, "rb") as pdf:
        reader = PyPDF2.PdfReader(pdf, strict=False)
        text = []

        for page in reader.pages:
            content = page.extract_text()
            text.append(content)

        return text

extractedText = extract('completed colleges/fortis.pdf')

book = load_workbook('data.xlsx')
sheet=book.worksheets[0]
classy = r"[A-Z]{3}\d{3} .+"
major = r"[A-Z]{3} \.+ .+"
majormatches = []
classymatches = []
df = pd.DataFrame(columns=['Colleges','Majors','Courses'])

medasst = "MAS  Medical Assisting"
medasstpat = r"MAS\d{3} .+"
medasstli = []
medofad = "MOA  Medical Office Administration"
medofadpat = r"MOA\d{3} .+"
medofadli = []

for ind in extractedText:
    common = re.findall(medasstpat, ind)
    if common:
        medasstli.extend(common)

for i in extractedText:
    com = re.findall(medofadpat, i)
    if com:
        medofadli.extend(com)


for tex in extractedText:
    mat = re.findall(major, tex)
    if mat:
        majormatches.extend(mat)
for text in extractedText:
    match = re.findall(classy, text)
    if match:
        classymatches.extend(match)


for mm in majormatches:
    if mm == "LTY    80" or mm == "EES    44":
        continue
    mm = mm.replace(".", "")
    mm = mm.replace("  ", " ")
    majorname = mm[0:3]
    if majorname == "PRM" or majorname == "PSY" or majorname == "LTY" or majorname == "EES":
        continue
    if majorname == "PHT":
        print('---------------------------------')
        print(medasst)
        for medclasses in medasstli:
            medclasses = medclasses.strip("  60 4")
            medclasses = medclasses.replace("  ", " ")
            medclasses = medclasses.strip(" 18")
            if medclasses == "MAS190 Externship":
                print(medclasses)
                df.loc[len(df.index)] = ["Fortis College-Smyrna",majorname, medclasses]

                break
            else:
                print(medclasses)
                df.loc[len(df.index)] = ["Fortis College-Smyrna",majorname, medclasses]

        print('---------------------------------')
        print(medofad)
        for mclass in medofadli:
            mclass = mclass.strip("  60 4")
            mclass = mclass.replace("  ", " ")
            print(mclass)
            df.loc[len(df.index)] = ["Fortis College-Smyrna",majorname, mclass]
    print('---------------------------------')
    print(mm)
    for m in classymatches:
        if len(m) <= 8:
            continue
        if m[0:3] == majorname:
            m = m.strip("  60 4")
            m = m.strip("    68")
            m = m.strip("    12")
            m = m.replace("  ", " ")
            print(m)
            df.loc[len(df.index)] = ["Fortis College-Smyrna",majorname, m]
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)
