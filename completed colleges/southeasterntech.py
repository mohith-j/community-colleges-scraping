from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl import load_workbook


url = "https://catalog.southeasterntech.edu/college-catalog/current/courses"
html = requests.get(url)
soup = BeautifulSoup(html.text, "html.parser")
classy = soup.findAll("div", attrs={"class":"page-title"})
book = load_workbook('data.xlsx')
sheet=book.worksheets[0]

df = pd.DataFrame(columns=['Colleges','Majors','Courses'])

li = []
for classes in classy:
    x = classes.text.replace("                                                - ", "")
    x = x.replace("\n", "")
    mjr = x[0:4]
    if "GUI" in mjr or "GEN" in mjr or "OGE" in mjr or "OCC" in mjr:
        continue
    if mjr not in li:
        print('---------------------------------')
        print(mjr)
        if "L&L" in x:
            final = x[8:]
            print(final)
            df.loc[len(df.index)] = ["Southeastern Technical College",mjr,final]
        else:
            final = x[9:]
            if " L" in final[0:2]:
                final = final[2:]
            print(final)
            df.loc[len(df.index)] = ["Southeastern Technical College",mjr,final]
        li.append(mjr)
    else:
        final = x[9:]
        if " L" in final[0:2]:
            final = final[2:]
        if final[0] == " ":
            final = final[1:]
        if "RNSG" in mjr and final[0] == "B":
            final = final[1:]
        print(final)
        df.loc[len(df.index)] = ["Southeastern Technical College",mjr,final]
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)
