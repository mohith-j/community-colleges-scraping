from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl import load_workbook


def get_classes(name, link):
    html=requests.get(link)
    soup=BeautifulSoup(html.text, "html.parser")
    classes=soup.findAll("li", attrs={"class":"acalog-course"})
    print(name)
    for clas in classes:
        clas_name=clas.find("a", attrs={"class":None}, href=True)
        print(clas_name.text)
        # insert writing spreadsheet code here
        df.loc[len(df.index)] = ["Georgia Perimeter College",name, clas_name.text]

book = load_workbook('data.xlsx')
sheet=book.worksheets[0]
df = pd.DataFrame(columns=['Colleges','Majors','Courses'])
url = "https://catalogs.gsu.edu/content.php?catoid=13&navoid=1599"
html = requests.get(url)
soup = BeautifulSoup(html.text, "html.parser")
majors = soup.findAll("li", attrs={"class":None})
c=0
for major in majors:
    if "African American Studies Pathway, A.A." in major.text:
        break
    c+=1
majors=majors[c:]
print(len(majors))
for major in majors:
    link=major.find("a", attrs={"class":None}, href=True)
    link="https://catalogs.gsu.edu/"+link["href"]
    # print(link)
    # print(major.text)
    get_classes(major.text, link)
    # link=major.find("a", attrs={"class":None}, href=True)
    
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)
    
