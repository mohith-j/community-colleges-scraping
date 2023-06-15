import PyPDF2
import re
import pandas as pd
from openpyxl import load_workbook

def extract(pdfs):
	with open(pdfs, "rb") as pdf:
		reader = PyPDF2.PdfReader(pdf, strict=False)
		text = []

		for page in reader.pages:
			content = page.extract_text()
			text.append(content)

		return text

extractedText = extract('completed colleges/laniertech.pdf')
book = load_workbook('data.xlsx')
sheet=book.worksheets[0]
df = pd.DataFrame(columns=['Colleges','Majors','Courses'])

pattern = r"[A-Z]{4} \d{4} [A-Z].+ \("
matches = []

for text in extractedText:
	match = re.findall(pattern, text)
	if match:
		matches.extend(match)


li = []
final = []
for m in matches:
	if m[0:9] not in li:
		li.append(m[0:9])
		final.append(m)
final = sorted(final)
currmjr = ""
for i in final:
	if currmjr != i[0:4]:
		print('---------------------------------')
		print(i[0:4])
	currmjr = i[0:4]
	if i[0:9] == "EMSP 2510":
		print("EMSP 2510 Clinical Applications for the Paramedic - I")
		df.loc[len(df.index)] = ["Lanier Technical College",currmjr,"EMSP 2510 Clinical Applications for the Paramedic - I"]
		continue
	if i[0:9] == "EMSP 2520":
		print("EMSP 2510 Clinical Applications for the Paramedic - II")
		df.loc[len(df.index)] = ["Lanier Technical College",currmjr,"EMSP 2510 Clinical Applications for the Paramedic - II"]
		continue
	if i[0:9] == "EMSP 2530":
		print("EMSP 2510 Clinical Applications for the Paramedic - III")
		df.loc[len(df.index)] = ["Lanier Technical College",currmjr,"EMSP 2510 Clinical Applications for the Paramedic - III"]
		continue
	i = i.strip("(")
	print(i)
	df.loc[len(df.index)] = ["Lanier Technical College",currmjr,i]
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)