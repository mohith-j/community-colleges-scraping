import PyPDF2
import re
import pandas as pd
from openpyxl import load_workbook


def extract(pdfs):
    with open(pdfs, "rb") as pdf:
        reader = PyPDF2.PdfReader(pdf, strict=False)
        text = []

        for page in reader.pages:
            content = page.extract_text()
            text.append(content)

        return text

extractedText = extract('completed colleges/piedmont.pdf')

book = load_workbook('data.xlsx')
sheet=book.worksheets[0]


df = pd.DataFrame(columns=['Colleges','Majors','Courses'])
classes = r"[A-Z]{4} \d{4} .+"
majors = r"[A-Z]{4} – .+"
majormatches = []
classmatches = []
#classmatches = sorted([*set(classmatches)])

for majo in extractedText:
    mat = re.findall(majors, majo)
    if mat:
        majormatches.extend(mat)

for clas in extractedText:
    match = re.findall(classes, clas)
    if match:
        classmatches.extend(match)

count = 0
#classmatches = sorted([*set(classmatches)])
res = []
final = []


for mm in majormatches:
    if "ACCT – Accounting" in mm or count > 0:
        count += 1
        print('---------------------------------')
        print(mm)
        mjr = mm[0:4]
        for x in classmatches:
            if x[0:10] not in res:
                res.append(x[0:10])
                final.append(x)
        for i in final:
            if i[0:4] == mjr:
                i = i.strip("                       3")
                i = i.strip("(4)")
                i = i.strip("(3)")
                i = i.strip("(2)")
                i = i.strip("(1)")
                i = i.strip(" (4)6")
                i = i.replace("     ", " ")
                i = i.replace("    ", " ")
                i = i.replace(" – ", " ")
                i = i.replace("  ", " ")
                i = i.replace("ALHS 1011 and ALHS 1090. The PN Dept. will take", "ALHS 1011 Structure and Function of the Human Body")
                i = i.replace("ALHS 1010 Introduction to Anatomy and Physiology", "")
                i = i.replace("ALHS 1090", "ALHS 1090 Medical Terminology for Allied Health Sciences")
                i = i.replace("BIOL 2117 Introduction t o Microbiology (3) + BIOL", "BIOL 2117 Introduction t o Microbiology")
                print(i)
                df.loc[len(df.index)] = ["Georgia Piedmont Technical College",mjr, i]                
    else:
        continue

with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)
