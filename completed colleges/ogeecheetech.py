from bs4 import BeautifulSoup
import requests
import csv
import pandas as pd
from openpyxl import load_workbook


def namechange(name):
    link = name.replace(" - ", "-")
    link = link.replace(" ", "-")
    link = "https://iq3.smartcatalogiq.com/Catalogs/Ogeechee-Technical-College/current/Catalog-and-Student-Handbook/Courses/"+link
    if "BUSN" in link:
        link = "https://iq3.smartcatalogiq.com/Catalogs/Ogeechee-Technical-College/current/Catalog-and-Student-Handbook/Courses/BUSN-Business-Technology"
    if "CWDS" in link:
        link = "https://iq3.smartcatalogiq.com/Catalogs/Ogeechee-Technical-College/current/Catalog-and-Student-Handbook/Courses/CWDS-Certified-Warehouse-Dist"
    if "ESCI" in link:
        link = "https://iq3.smartcatalogiq.com/Catalogs/Ogeechee-Technical-College/current/Catalog-and-Student-Handbook/Courses/ESCI-Enviornmental-Technology"
    if "FWMT" in link:
        link = "https://iq3.smartcatalogiq.com/Catalogs/Ogeechee-Technical-College/current/Catalog-and-Student-Handbook/Courses/FWMT-Forestry-Wildlife-Mgmt"
    r = requests.get(link)
    if r.status_code != 404:
        return link
    else:
        return None

df = pd.DataFrame(columns=['Colleges','Majors','Courses'])
url = "https://iq3.smartcatalogiq.com/en/Catalogs/Ogeechee-Technical-College/current/Catalog-and-Student-Handbook/Courses"
html = requests.get(url)
soup = BeautifulSoup(html.text, "html.parser")
majors = soup.findAll("a", attrs={"class": None}, href=True)
book = load_workbook('data.xlsx')
sheet=book.worksheets[0]


for major in majors:
    if "2022-2023 Catalog & Student Handbook" in major.text:
        continue
    if "-" in major.text:
        print('---------------------------------')
        print(major.text)
        namechange(major.text)
        newurl = namechange(major.text)
        newhtml = requests.get(newurl)
        newsoup = BeautifulSoup(newhtml.text, "html.parser")
        classes = newsoup.findAll("a", attrs={"class": None}, href=True)
        for classy in classes:
            if "2022-2023 Catalog & Student Handbook" in classy.text:
                continue
            if ("0" in classy.text or "1" in classy.text or "2" in classy.text or "3" in classy.text) and len(classy.text) > 5:
                print(classy.text)
                df.loc[len(df.index)] = ["Ogeechee Technical College",major.text, classy.text]
        if "Welding" in major.text:
            break
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)