from bs4 import BeautifulSoup
import requests
import csv
import pandas as pd
from openpyxl import load_workbook

li = []
count = 1

df = pd.DataFrame(columns=['Colleges','Majors','Courses'])
book = load_workbook('data.xlsx')
sheet=book.worksheets[0]

for i in range(1,9):
    if count == 1:
        url = "http://catalog.gwinnetttech.edu/content.php?catoid=3&catoid=3&navoid=117&filter%5Bitem_type%5D=3&filter%5Bonly_active%5D=1&filter%5B3%5D=1&filter%5Bcpage%5D=1#acalog_template_course_filter"
        count += 1
    elif count == 2:
        url = "http://catalog.gwinnetttech.edu/content.php?catoid=3&catoid=3&navoid=117&filter%5Bitem_type%5D=3&filter%5Bonly_active%5D=1&filter%5B3%5D=1&filter%5Bcpage%5D=2#acalog_template_course_filter"
        count += 1
    elif count == 3:
        url = "http://catalog.gwinnetttech.edu/content.php?catoid=3&catoid=3&navoid=117&filter%5Bitem_type%5D=3&filter%5Bonly_active%5D=1&filter%5B3%5D=1&filter%5Bcpage%5D=3#acalog_template_course_filter"
        count += 1
    elif count == 4:
        url = "http://catalog.gwinnetttech.edu/content.php?catoid=3&catoid=3&navoid=117&filter%5Bitem_type%5D=3&filter%5Bonly_active%5D=1&filter%5B3%5D=1&filter%5Bcpage%5D=4#acalog_template_course_filter"
        count += 1
    elif count == 5:
        url = "http://catalog.gwinnetttech.edu/content.php?catoid=3&catoid=3&navoid=117&filter%5Bitem_type%5D=3&filter%5Bonly_active%5D=1&filter%5B3%5D=1&filter%5Bcpage%5D=5#acalog_template_course_filter"
        count += 1
    elif count == 6:
        url = "http://catalog.gwinnetttech.edu/content.php?catoid=3&catoid=3&navoid=117&filter%5Bitem_type%5D=3&filter%5Bonly_active%5D=1&filter%5B3%5D=1&filter%5Bcpage%5D=6#acalog_template_course_filter"
        count += 1
    elif count == 7:
        url = "http://catalog.gwinnetttech.edu/content.php?catoid=3&catoid=3&navoid=117&filter%5Bitem_type%5D=3&filter%5Bonly_active%5D=1&filter%5B3%5D=1&filter%5Bcpage%5D=7#acalog_template_course_filter"
        count += 1
    elif count == 8:
        url = "http://catalog.gwinnetttech.edu/content.php?catoid=3&catoid=3&navoid=117&filter%5Bitem_type%5D=3&filter%5Bonly_active%5D=1&filter%5B3%5D=1&filter%5Bcpage%5D=8#acalog_template_course_filter"

    html = requests.get(url)
    soup = BeautifulSoup(html.text, "html.parser")
    classy = soup.findAll("a", attrs={"class":None}, href=True)




    for classes in classy:
        if "-" not in classes.text or "70-" in classes.text or "Drug" in classes.text or "Alph" in classes.text:
            continue
        mjr = classes.text[0:4]
        if mjr not in li:
            print('---------------------------------')
            print(mjr)
            li.append(mjr)
            final = classes.text[12:]
            if final[0] == " ":
                final = final[1:]
            print(final)
        else:
            if classes.text[9] == "L":
                final = classes.text[13:]
            else:
                final = classes.text[12:]
            if final[0] == " ":
                final = final[1:]
            print(final)
            df.loc[len(df.index)] = ["Gwinnett Technical College",mjr, classes.text]
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)

