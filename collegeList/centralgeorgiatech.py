import PyPDF2
import re
import pandas as pd
from openpyxl import load_workbook

def extract(pdfs):
    with open(pdfs, "rb") as pdf:
        reader = PyPDF2.PdfReader(pdf, strict=False)
        text = []

        for page in reader.pages:
            content = page.extract_text()
            text.append(content)

        return text
#AC-Catalog-2022-2023-08082022
extractedText = extract('collegeList/centralgeorgiatech.pdf')

book = load_workbook('data.xlsx')
sheet=book.worksheets[0]

df = pd.DataFrame(columns=['Colleges','Majors','Courses'])
pattern = r"[A-Z]{4} \d{4} \| .+"
matches = []

for text in extractedText:
    match = re.findall(pattern, text)
    if match:
        matches.extend(match)

count = 0
currmjr = ""
for m in matches:
    if currmjr != m[0:4]:
        print('---------------------------------')
        print(m[0:4])
    currmjr = m[0:4]
    m = re.sub("\(.*?\)","",m)
    m = re.sub("\(.*?\-", "", m)
    print(m)
    df.loc[len(df.index)] = ["Central Georgia Technical College",currmjr, m]
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)