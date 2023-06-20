import PyPDF2
import re
import pandas as pd
from openpyxl import load_workbook

def extract(pdfs):
	with open(pdfs, "rb") as pdf:
		reader = PyPDF2.PdfReader(pdf, strict=False)
		text = []

		for page in reader.pages:
			content = page.extract_text()
			text.append(content)

		return text

extractedText = extract('collegeList/atlantainstmusmed.pdf')

df = pd.DataFrame(columns=['Colleges','Majors','Courses'])
pattern = r"[A-Z]{3}\d{3} [A-Z].+"
matches = []
book = load_workbook('data.xlsx')
sheet=book.worksheets[0]

for text in extractedText:
	match = re.findall(pattern, text)
	if match:
		matches.extend(match)


li = []
final = []
for m in matches:
	m = m.strip("    1")
	m = m.strip("*")
	m = m.strip("*     2")
	if "*    3Associate of Applied Science in Music and Technology — Total Credit Hours: 94COURSES" in m:
		m = m.strip("*    3Associate of Applied Science in Music and Technology — Total Credit Hours: 94COURSES")
	m = m.strip("*     3")
	m = m.strip("*     4")
	m = m.strip("    5")
	m = m.strip("    6")
	if m == "110":
		continue
	if m[0:6] not in li:
		li.append(m[0:6])
		final.append(m)
final = sorted(final)
currmjr = ""
for i in final:
	if currmjr != i[0:3]:
		print('---------------------------------')
		print(i[0:3])
	currmjr = i[0:3]
	print(i)
	df.loc[len(df.index)] = ["Atlanta Institute of Music and Media",currmjr, i]
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)