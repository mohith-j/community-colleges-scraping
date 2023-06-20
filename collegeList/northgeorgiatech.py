import PyPDF2
import re
import pandas as pd
from openpyxl import load_workbook

def extract(pdfs):
    with open(pdfs, "rb") as pdf:
        reader = PyPDF2.PdfReader(pdf, strict=False)
        text = []

        for page in reader.pages:
            content = page.extract_text()
            text.append(content)

        return text

extractedText = extract('collegeList/northgeorgiatech.pdf')
book = load_workbook('data.xlsx')
sheet=book.worksheets[0]
df = pd.DataFrame(columns=['Colleges','Majors','Courses'])

pattern = r"[A-Z]{4} \d{4} – [A-Z].+ \("
matches = []

for text in extractedText:
    match = re.findall(pattern, text)
    if match:
        matches.extend(match)


count = 0
matches = sorted(matches)
currmjr = ""
for m in matches:
    if currmjr != m[0:4]:
        print('---------------------------------')
        print(m[0:4])
    currmjr = m[0:4]
    m = m.strip("(")
    m = m.replace("(3-0-3)", "")
    m = m.replace("(4-0-4)", "")
    m = m.replace("(2-3-3)", "")
    m = m.replace("(Basic Skills – non-degree level)Prerequisites: ENGL 0090 or Appropriate Writing", "")
    m = m.replace("(4-2-5)Prerequisites (diploma): ALHS 1011, ALHS 1090, ENGL 1010, MATH 1012Prerequisites", "")
    print(m)
    df.loc[len(df.index)] = ["North Georgia Technical College",currmjr,m]
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)