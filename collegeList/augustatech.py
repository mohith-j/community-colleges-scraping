from bs4 import BeautifulSoup
import requests
import csv
import pandas as pd
from openpyxl import load_workbook


def namechange(name):
	link = name.replace(" & ", "-")
	link = link.replace(" - ","-")
	link = link.replace("- ", "-")
	link = link.replace("/", "-")
	link = link.replace(".", "")
	link = link.replace(" ", "-").lower()
	link="https://augustatech.smartcatalogiq.com/en/2023/semester-catalog/course-descriptions/"+link
	r = requests.get(link)
	if r.status_code != 404:
		return link
	else:
		return None



df = pd.DataFrame(columns=['Colleges','Majors','Courses'])
url = "https://augustatech.smartcatalogiq.com/en/2023/semester-catalog/course-descriptions/"
html = requests.get(url)
soup = BeautifulSoup(html.text, "html.parser")
majors = soup.findAll("a", attrs={"class":None}, href=True)
book = load_workbook('data.xlsx')
sheet=book.worksheets[0]


for major in majors:
    if "-" in major.text:
        print('---------------------------------')
        print(major.text)
        newurl = namechange(major.text)
        newhtml = requests.get(newurl)
        newsoup = BeautifulSoup(newhtml.text, "html.parser")
        classes = newsoup.findAll("a", attrs={"class":None}, href=True)
        for classy in classes:
            if "Summer Semester 2023 Catalog" in classy.text:
                continue
            if classy.text == "1000" or classy.text == "2000":
                continue
            if "1" in classy.text or "2" in classy.text or "3" in classy.text:
                print(classy.text)
                df.loc[len(df.index)] = ["Augusta Technical College",major.text, classy.text]
    if "Welding" in major.text:
        break
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)


