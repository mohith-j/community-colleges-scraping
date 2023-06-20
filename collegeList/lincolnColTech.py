from bs4 import BeautifulSoup
import requests
import csv
import pandas as pd
from openpyxl import load_workbook



def namechange(name):
	link=name.replace(" - ","-")
	link = link.replace("- ", "-")
	link = link.replace("/", "-")
	link =link.replace(" ", "-")
	link="https://chattahoocheetech.smartcatalogiq.com/en/2022-2023/General-Catalog/Courses/"+link
	if "CMTT" in name:
		link="https://chattahoocheetech.smartcatalogiq.com/2022-2023/General-Catalog/Courses/CMMT"
	if "DMPT" in name:
		link = "https://chattahoocheetech.smartcatalogiq.com/2022-2023/General-Catalog/Courses/DMPT-Design-and-Media-Production"
	if "ELUT" in name:
		link = "https://chattahoocheetech.smartcatalogiq.com/2022-2023/General-Catalog/Courses/ELUT"
	r = requests.get(link)
	if r.status_code != 404:
		return link
	else:
		return None


df = pd.DataFrame(columns=['Colleges','Majors','Courses'])
firsturl = "https://www.lincolntech.edu/careers/skilled-trades/hvac/air-conditioning-refrigeration-heating-technology-5"
firsthtml = requests.get(firsturl)
firstsoup = BeautifulSoup(firsthtml.text, "html.parser")
firstmajors = firstsoup.findAll("h1", attrs={"class": None})
book = load_workbook('data.xlsx')
sheet=book.worksheets[0]

for firstmajor in firstmajors:
	fir = firstmajor.text.replace(" — Marietta", "")
	print('---------------------------------')
	print(fir)
	firstclasses = firstsoup.findAll("span", attrs={"property": "schema:name"})
	for firstclass in firstclasses:
		print(firstclass.text)
		df.loc[len(df.index)] = ["Lincoln College of Technology",fir, firstclass.text]

securl = "https://www.lincolntech.edu/careers/skilled-trades/electrical/electrical-and-electronic-systems-technician-1"
sechtml = requests.get(securl)
secsoup = BeautifulSoup(sechtml.text, "html.parser")
secmajors = secsoup.findAll("h1", attrs={"class": None})
for secmajor in secmajors:
	sec = secmajor.text.replace(" — Marietta", "")
	print('---------------------------------')
	print(sec)
	secclasses = secsoup.findAll("span", attrs={"property": "schema:name"})
	for secclass in secclasses:
		if secclass.text[0] == " ":
			last = secclass.text[1:]
			print(last)
			df.loc[len(df.index)] = ["Lincoln College of Technology",sec, last]
		else:
			print(secclass.text)
			df.loc[len(df.index)] = ["Lincoln College of Technology",sec, secclass.text]

thiurl = "https://www.lincolntech.edu/careers/skilled-trades/electrical/electrical-and-electronic-systems-technician-2"
thihtml = requests.get(thiurl)
thisoup = BeautifulSoup(thihtml.text, "html.parser")
thimajors = thisoup.findAll("h1", attrs={"class": None})
for thimajor in thimajors:
	thi = thimajor.text.replace(" — Marietta", "")
	print('---------------------------------')
	print(thi)
	thiclasses = thisoup.findAll("span", attrs={"property": "schema:name"})
	for thiclass in thiclasses:
		if thiclass.text[0] == " ":
			last = thiclass.text[1:]
			print(last)
			df.loc[len(df.index)] = ["Lincoln College of Technology",thi, last]
		else:
			print(thiclass.text)
			df.loc[len(df.index)] = ["Lincoln College of Technology",thi, thiclass.text]


foururl = "https://www.lincolntech.edu/careers/health-sciences/medical-assistant-technology/medical-assistant-technology-courses-4"
fourhtml = requests.get(foururl)
foursoup = BeautifulSoup(fourhtml.text, "html.parser")
fourmajors = foursoup.findAll("h1", attrs={"class": None})
for fourmajor in fourmajors:
	four = fourmajor.text.replace(" — Marietta", "")
	print('---------------------------------')
	print(four)
	fourclasses = foursoup.findAll("span", attrs={"property": "schema:name"})
	for fourclass in fourclasses:
		if fourclass.text[0] == " ":
			last = fourclass.text[1:]
			print(last)
			df.loc[len(df.index)] = ["Lincoln College of Technology",four, last]
		else:
			print(fourclass.text)
			df.loc[len(df.index)] = ["Lincoln College of Technology",four, fourclass.text]


fiveurl = "https://www.lincolntech.edu/careers/health-sciences/medical-assistant/medical-assistant-10"
fivehtml = requests.get(fiveurl)
fivesoup = BeautifulSoup(fivehtml.text, "html.parser")
fivemajors = fivesoup.findAll("h1", attrs={"class": None})
for fivemajor in fivemajors:
	five = fivemajor.text.replace(" — Marietta", "")
	print('---------------------------------')
	print(five)
	fiveclasses = fivesoup.findAll("span", attrs={"property": "schema:name"})
	for fiveclass in fiveclasses:
		if fiveclass.text[0] == " ":
			last = fiveclass.text[1:]
			print(last)
			df.loc[len(df.index)] = ["Lincoln College of Technology",five, last]
		else:
			print(fiveclass.text)
			df.loc[len(df.index)] = ["Lincoln College of Technology",five, fiveclass.text]
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)
