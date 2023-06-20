from bs4 import BeautifulSoup
import requests
import csv
import pandas as pd
from openpyxl import load_workbook

def namechange(name):
	link = name.replace(" - ","-")
	link = link.replace("- ", "-")
	link = link.replace("/", "-")
	link = link.replace(" ", "-").lower()
	link = "https://gntc.smartcatalogiq.com/en/2023-2024/semester-catalog/course-descriptions/"+link+"/"
	if "busn" in link:
		link = "https://gntc.smartcatalogiq.com/en/2023-2024/semester-catalog/course-descriptions/busn-business-administrative-techno/"
	if "flpd" in link:
		link = "https://gntc.smartcatalogiq.com/en/2023-2024/semester-catalog/course-descriptions/flpd-flooring-production/"
	if "lact" in link:
		link = "https://gntc.smartcatalogiq.com/en/2023-2024/semester-catalog/course-descriptions/lact-lactation/"
	if "ped" in link:
		link = "https://gntc.smartcatalogiq.com/en/2023-2024/semester-catalog/course-descriptions/peds-pediatric-echocardiography/"
	r = requests.get(link)
	if r.status_code != 404:
		return link
	else:
		return None

df = pd.DataFrame(columns=['Colleges','Majors','Courses'])
url = "https://gntc.smartcatalogiq.com/en/2023-2024/semester-catalog/course-descriptions/"
html = requests.get(url)
soup = BeautifulSoup(html.text, "html.parser")
majors = soup.findAll("a", attrs={"class":None}, href=True)
book = load_workbook('data.xlsx')
sheet=book.worksheets[0]


for major in majors:
	if "General Catalog" in major.text or "2023-2024 Catalog" in major.text or "Programs of Study" in major.text or "Programs-of-Study" in major.text:
		continue
	if "-" in major.text:
		print('---------------------------------')
		print(major.text)
		newurl = namechange(major.text)
		newhtml = requests.get(newurl)
		newsoup = BeautifulSoup(newhtml.text, "html.parser")
		classes = newsoup.findAll("a", attrs={"class":None}, href=True)
		for classy in classes:
			if "General Catalog" in classy.text or "2023-2024 Catalog" in classy.text or "Programs Available 100% Online" in classy.text:
				continue
			if ("0" in classy.text or "1" in classy.text or "2" in classy.text) and len(classy.text) > 5:
				print(classy.text)
				df.loc[len(df.index)] = ["Georgia Northwestern Technical College",major.text, classy.text]
		if "Welding" in major.text:
			break
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)
