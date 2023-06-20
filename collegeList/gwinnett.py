from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl import load_workbook

def get_classes(name, link):
    html=requests.get(link)
    soup=BeautifulSoup(html.text, "html.parser")
    classes=soup.findAll("li", attrs={"class":None})
    print("Program Name:"+name)
    for clas in classes:
        print(clas.text)
        # insert writing spreadsheet code here
        df.loc[len(df.index)] = ["Gwinnett College-Lilburn",name, clas.text]
book = load_workbook('data.xlsx')
sheet=book.worksheets[0]
df = pd.DataFrame(columns=['Colleges','Majors','Courses'])
url = "https://www.gwinnettcollege.edu/locations/lilburn/"
html = requests.get(url)
soup = BeautifulSoup(html.text, "html.parser")
div = soup.find("div", attrs={"class":"entry-content"})
majors=div.findAll("a", attrs={"class":None})
for major in majors:
    # print(major["href"])
    # print(major.text)
    get_classes(major.text, major["href"])
    
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)
