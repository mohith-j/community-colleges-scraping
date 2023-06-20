from bs4 import BeautifulSoup
import requests
import csv
import pandas as pd
from openpyxl import load_workbook


def namechange(name):
	link = name.replace(" - ","-")
	link = link.replace("- ", "-")
	link = link.replace("/", "-")
	link = link.replace(" ", "-").lower()
	link="https://iq2.smartcatalogiq.com/en/Catalogs/Columbus-Technical-College/2021-2022/2021-22-Catalog-and-Student-Handbook/Courses/"+link
	if "cism" in link:
		link = "https://westgatech.smartcatalogiq.com/en/2022-2023/student-catalog/course-descriptions/cism-computer-information-systems/"
	if "gert" in link:
		link = "https://westgatech.smartcatalogiq.com/en/2022-2023/student-catalog/course-descriptions/copy-of-nast-nurse-aide/"
	if "phar" in link:
		link = "https://westgatech.smartcatalogiq.com/en/2022-2023/student-catalog/course-descriptions/phar-pharmacy-assistant/"
	if "ecet" in link:
		link = "https://iq2.smartcatalogiq.com/Catalogs/Columbus-Technical-College/2021-2022/2021-22-Catalog-and-Student-Handbook/Courses/ECET-Electrical-Computer-Engineering-Tech"
	if "mcht" in link:
		link = "https://iq2.smartcatalogiq.com/en/Catalogs/Columbus-Technical-College/2021-2022/2021-22-Catalog-and-Student-Handbook/Courses/MCHT-Precision-Machining-and-Manufacturing-Technology"
	r = requests.get(link)
	if r.status_code != 404:
		return link
	else:
		return None


df = pd.DataFrame(columns=['Colleges','Majors','Courses'])
url = "https://iq2.smartcatalogiq.com/en/Catalogs/Columbus-Technical-College/2021-2022/2021-22-Catalog-and-Student-Handbook/Courses"
html = requests.get(url)
soup = BeautifulSoup(html.text, "html.parser")
majors = soup.findAll("a", attrs={"class":None}, href=True)
book = load_workbook('data.xlsx')
sheet=book.worksheets[0]


for major in majors:
	if "General Catalog" in major.text or "2021-22 Catalog and Student Handbook" in major.text:
		continue
	if "-" in major.text:
		print('---------------------------------')
		print(major.text)
		newurl = namechange(major.text)
		newhtml = requests.get(newurl)
		newsoup = BeautifulSoup(newhtml.text, "html.parser")
		classes = newsoup.findAll("a", attrs={"class":None}, href=True)
		for classy in classes:
			if "General Catalog" in classy.text or "SEMN 1000" in classy.text or "2021-22 Catalog and Student Handbook" in classy.text:
				continue
			if ("0" in classy.text or "1" in classy.text or "2" in classy.text or "3" in classy.text) and len(classy.text) > 5:
				if (classy.text[0] == " " and classy.text[1] == " "):
					last = classy.text[2:]
					print(last)
					df.loc[len(df.index)] = ["Columbus Technical College",major.text, last]
				elif classy.text[0] == " ":
					last = classy.text[1:]
					print(last)
					df.loc[len(df.index)] = ["Columbus Technical College",major.text, last]
				else:
					print(classy.text)
					df.loc[len(df.index)] = ["Columbus Technical College",major.text, classy.text]
		if major.text == "WELD - Welding":
			break
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)
 