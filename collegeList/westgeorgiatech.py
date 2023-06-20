from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl import load_workbook

def namechange(name):
	link = name.replace(" - ","-")
	link = link.replace("- ", "-")
	link = link.replace("/", "-")
	link = link.replace(" ", "-").lower()
	link="https://westgatech.smartcatalogiq.com/en/2022-2023/student-catalog/course-descriptions/"+link
	if "cism" in link:
		link = "https://westgatech.smartcatalogiq.com/en/2022-2023/student-catalog/course-descriptions/cism-computer-information-systems/"
	if "gert" in link:
		link = "https://westgatech.smartcatalogiq.com/en/2022-2023/student-catalog/course-descriptions/copy-of-nast-nurse-aide/"
	if "phar" in link:
		link = "https://westgatech.smartcatalogiq.com/en/2022-2023/student-catalog/course-descriptions/phar-pharmacy-assistant/"
	r = requests.get(link)
	if r.status_code != 404:
		return link
	else:
		return None


url = "https://westgatech.smartcatalogiq.com/en/2022-2023/student-catalog/course-descriptions/"
html = requests.get(url)
soup = BeautifulSoup(html.text, "html.parser")
majors = soup.findAll("a", attrs={"class":None}, href=True)
book = load_workbook('data.xlsx')
sheet=book.worksheets[0]

df = pd.DataFrame(columns=['Colleges','Majors','Courses'])

for major in majors:
	if "General Catalog" in major.text or "Student Catalog and Handbook 2022-2023" in major.text:
		continue
	if "-" in major.text:
		print('---------------------------------')
		print(major.text)
		newurl = namechange(major.text)
		newhtml = requests.get(newurl)
		newsoup = BeautifulSoup(newhtml.text, "html.parser")
		classes = newsoup.findAll("a", attrs={"class":None}, href=True)
		for classy in classes:
			if "General Catalog" in classy.text or "Student Catalog and Handbook 2022-2023" in classy.text:
				continue
			if ("0" in classy.text or "1" in classy.text or "2" in classy.text) and len(classy.text) > 5:
				print(classy.text)
				df.loc[len(df.index)] = ["West Georgia Technical College",major.text, classy.text]
		if "Welding" in major.text:
			break
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)
