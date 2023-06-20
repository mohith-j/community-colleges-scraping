from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl import load_workbook

url = "https://oftc.smartcatalogiq.com/en/2022-2023/ay23-academic-catalog-handbook/courses/"
html = requests.get(url)
soup = BeautifulSoup(html.text, "html.parser")
classes = soup.findAll("a", attrs={"class":None}, href=True)
book = load_workbook('data.xlsx')
sheet=book.worksheets[0]

df = pd.DataFrame(columns=['Colleges','Majors','Courses'])

count = 0
currmjr = ""
for classy in classes:
    if "XXXX XXXX" in classy.text or "  Choose 3 or more credit hours:" in classy.text:
        continue
    if "0" in classy.text or "1" in classy.text or "2" in classy.text or "3" in classy.text:
        if currmjr != classy.text[0:4]:
            print('---------------------------------')
            print(classy.text[0:4])
        currmjr = classy.text[0:4]
        print(classy.text)
        df.loc[len(df.index)] = ["Oconee Fall Line Technical College",currmjr,classy.text]
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)

