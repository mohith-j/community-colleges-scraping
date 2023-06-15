import PyPDF2
import re
import pandas as pd
from openpyxl import load_workbook

def extract(pdfs):
    with open(pdfs, "rb") as pdf:
        reader = PyPDF2.PdfReader(pdf, strict=False)
        text = []

        for page in reader.pages:
            content = page.extract_text()
            text.append(content)

        return text
#AC-Catalog-2022-2023-08082022
extractedText = extract('completed colleges/guptonjones.pdf')

book = load_workbook('data.xlsx')
sheet=book.worksheets[0]



df = pd.DataFrame(columns=['Colleges','Majors','Courses'])
pattern = r"[A-Z]{3} [0-9]{3} [A-Z].+\("
matches = []

for text in extractedText:
    match = re.findall(pattern, text)
    if match:
        matches.extend(match)

#matches = sorted(matches)
for m in matches:
    m=m[:-1]
    print(m)
    df.loc[len(df.index)] = ["Andrew College",MAJORNAME, m]
with pd.ExcelWriter('data.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df.to_excel(writer,sheet_name="Sheet",header=False, index=False, startrow=sheet.max_row)
